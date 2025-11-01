#!/usr/bin/env python
# coding: utf-8

# In[1]:


import os
import sys
import json
import pandas as pd
from glob import glob

# Optional imports for embedding images into Excel
try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.utils import get_column_letter
except Exception:
    load_workbook = None
    OpenpyxlImage = None
    get_column_letter = None

try:
    from PIL import Image as PILImage
except Exception:
    PILImage = None

def extract_messages_data(folder_path):
    data = []
    group_folders = glob(os.path.join(folder_path, 'Google Chat', 'Groups', '*'))

    for group_folder in group_folders:
        group_type = "DM" if "DM" in group_folder else "Space"
        group_id = os.path.basename(group_folder).split()[-1]

        messages_path = os.path.join(group_folder, 'messages.json')
        group_info_path = os.path.join(group_folder, 'group_info.json')

        # Extract space name from group_info.json if it's a "Space" type
        space_name = None
        if group_type == "Space" and os.path.exists(group_info_path):
            with open(group_info_path, 'r', encoding='utf-8') as group_info_file:
                group_info = json.load(group_info_file)
                space_name = group_info.get("name")

        # Helper to discover image attachment file paths referenced in a message
        def find_attachment_paths(msg_dict):
            image_exts = ('.png', '.jpg', '.jpeg', '.gif', '.webp', '.bmp', '.svg')
            paths = []

            def walk(obj):
                if isinstance(obj, dict):
                    for v in obj.values():
                        walk(v)
                elif isinstance(obj, list):
                    for item in obj:
                        walk(item)
                elif isinstance(obj, str):
                    lower = obj.lower()
                    # if the string looks like a filename with an image extension
                    if any(lower.endswith(ext) or (ext in lower and '/' in obj) for ext in image_exts):
                        # Try a few candidate locations relative to group_folder
                        candidates = [
                            os.path.join(group_folder, obj),
                            os.path.join(group_folder, os.path.basename(obj)),
                        ]
                        # Also allow glob search within the group folder
                        for cand in candidates:
                            if os.path.exists(cand):
                                paths.append(os.path.abspath(cand))
                                break
                        else:
                            # glob search for basename
                            matches = glob(os.path.join(group_folder, '**', os.path.basename(obj)), recursive=True)
                            for m in matches:
                                if os.path.exists(m):
                                    paths.append(os.path.abspath(m))
                                    break

            walk(msg_dict)
            # deduplicate while preserving order
            seen = set()
            out = []
            for p in paths:
                if p not in seen:
                    seen.add(p)
                    out.append(p)
            return out

        # Process each message in messages.json
        if os.path.exists(messages_path):
            with open(messages_path, 'r', encoding='utf-8') as messages_file:
                messages_json = json.load(messages_file)
                for message in messages_json.get("messages", []):
                    attachment_paths = find_attachment_paths(message)
                    record = {
                        "GroupType": group_type,
                        "GroupId": group_id,
                        "SendersName": message["creator"].get("name"),
                        "SendersEmail": message["creator"].get("email"),
                        "MessageTimestamp": message.get("created_date"),
                        "MessageContent": message.get("text"),
                        "SpaceName": space_name if group_type == "Space" else None,
                        "AttachmentPaths": attachment_paths,
                        "AttachmentNames": '; '.join([os.path.basename(p) for p in attachment_paths]) if attachment_paths else ''
                    }
                    data.append(record)

    return data

# Prompt the user for the TakeOut folder location (default: ./TakeOut)
current_dir = os.getcwd()
default_folder = os.path.join(current_dir, 'TakeOut')
try:
    user_input = input(f"Enter path to TakeOut folder [default: {default_folder}]: ").strip()
except Exception:
    # In non-interactive environments where input() may fail, fall back to default
    user_input = ''

folder_path = os.path.expanduser(user_input) if user_input else default_folder
folder_path = os.path.abspath(folder_path)

if not os.path.exists(folder_path) or not os.path.isdir(folder_path):
    print(f"Error: The path '{folder_path}' does not exist or is not a directory.")
    sys.exit(1)

# Extract the data and load it into a DataFrame
messages_data = extract_messages_data(folder_path)
df = pd.DataFrame(messages_data)

# Prepare DataFrame for saving (AttachmentPaths is a list; keep AttachmentNames for display)
df_to_save = df.copy()
if 'AttachmentPaths' in df_to_save.columns:
    try:
        df_to_save = df_to_save.drop(columns=['AttachmentPaths'])
    except Exception:
        pass

# Save the DataFrame to an Excel file inside the provided TakeOut folder
output_path = os.path.join(folder_path, 'GoogleChatMessages.xlsx')
df_to_save.to_excel(output_path, index=False)
print(f"Data saved to {output_path}")

# Try to embed images into the spreadsheet if openpyxl and pillow are available
if load_workbook is None or OpenpyxlImage is None:
    print("Note: openpyxl not available — images will not be embedded. Install with: pip install openpyxl pillow")
else:
    try:
        wb = load_workbook(output_path)
        ws = wb.active

        # Map headers to column indices
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        header_map = {cell.value: idx + 1 for idx, cell in enumerate(header_row)}

        # Decide where to put images: insert columns after MessageContent (or append)
        message_col = header_map.get('MessageContent')

        # Determine the maximum number of attachment images across all messages
        try:
            max_images = max((len(m.get('AttachmentPaths', [])) for m in messages_data), default=0)
        except Exception:
            max_images = 0

        if max_images <= 0:
            # nothing to embed
            print('No image attachments found to embed.')
            wb.close()

        if message_col is None:
            image_col_idx = ws.max_column + 1
            # append columns for each image
            for i in range(max_images):
                ws.cell(row=1, column=image_col_idx + i, value=f'AttachmentImage{i+1}')
        else:
            image_col_idx = message_col + 1
            # insert required number of columns at once
            ws.insert_cols(image_col_idx, amount=max_images)
            for i in range(max_images):
                ws.cell(row=1, column=image_col_idx + i, value=f'AttachmentImage{i+1}')

        import tempfile
        temp_files = []

        # Iterate over DataFrame rows and embed the first attachment image (if any)
        for idx, row in df.iterrows():
            excel_row = idx + 2  # account for header
            attachment_paths = None
            try:
                attachment_paths = row['AttachmentPaths'] if 'AttachmentPaths' in row.index else None
            except Exception:
                attachment_paths = None

            # Fallback to messages_data if necessary
            if attachment_paths is None:
                try:
                    attachment_paths = messages_data[idx].get('AttachmentPaths', [])
                except Exception:
                    attachment_paths = []

            if not attachment_paths:
                continue

            # Embed each attachment into its own column (up to max_images)
            for j in range(min(len(attachment_paths), max_images)):
                ap = attachment_paths[j]
                if not ap or not os.path.exists(ap):
                    continue

                try:
                    img_path_to_use = ap
                    # Resize using PIL to a reasonable thumbnail size to avoid huge embedded images
                    if PILImage is not None:
                        try:
                            pil_img = PILImage.open(ap)
                            max_size = (150, 150)
                            pil_img.thumbnail(max_size)
                            ext = os.path.splitext(ap)[1].lower() or '.png'
                            tf = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
                            fmt = pil_img.format if hasattr(pil_img, 'format') and pil_img.format else 'PNG'
                            pil_img.save(tf, format=fmt)
                            tf.close()
                            img_path_to_use = tf.name
                            temp_files.append(tf.name)
                        except Exception:
                            img_path_to_use = ap

                    img = OpenpyxlImage(img_path_to_use)
                    target_col = image_col_idx + j
                    cell = ws.cell(row=excel_row, column=target_col)
                    ws.add_image(img, cell.coordinate)
                    # Adjust row height and column width so the image is visible
                    try:
                        ws.row_dimensions[excel_row].height = 110
                        if get_column_letter is not None:
                            col_letter = get_column_letter(target_col)
                            # set column width to accommodate thumbnail
                            ws.column_dimensions[col_letter].width = 20
                    except Exception:
                        pass
                except Exception:
                    # don't let image embedding break the whole script
                    continue

        # Save workbook and cleanup temp files
        wb.save(output_path)
        for t in temp_files:
            try:
                os.remove(t)
            except Exception:
                pass
        print(f"Images embedded (where available) into {output_path}")
    except Exception as e:
        print(f"Warning: failed to embed images into Excel: {e}")


# In[ ]:




